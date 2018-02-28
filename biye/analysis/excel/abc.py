# !/usr/bin/env python3
# -*- coding: utf-8 -*-


import openpyxl
import pymysql
import datetime

host = 'localhost'
user = 'root'
pwd = ''
db = 'tieba'
sql = 'select * from pure_users'
sheet_name = 'pure_users'
out_path = datetime.datetime.now().strftime('%Y%m%d') + '.xlsx'

conn = pymysql.connect(host, user, pwd, db, charset='utf8')
cursor = conn.cursor()
count = cursor.execute(sql)
print(count)

cursor.scroll(0, mode='absolute')
results = cursor.fetchall()
fields = cursor.description

title = []
for name in fields:
    title.append(name[0])

outwb = openpyxl.Workbook()
ws1 = outwb.active
ws1.title = "res"

out_filename = r'result.xlsx'

# 添加标题
for x in range(0, len(title)):
    ws1.cell(column=x+1, row=1, value="%s" % title[x])

# 添加内容
i = 2
for data_l in results:
    for x in range(0, len(data_l)):
        ws1.cell(column=x + 1, row=i, value="%s" % data_l[x])
    i += 1

outwb.save(filename=out_filename)
